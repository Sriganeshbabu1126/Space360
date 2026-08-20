import csv
import io
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from typing import List, Dict, Any

# Map roles to fields
ADMIN_FIELDS = [
    "id", "title", "status", "priority", "site", "assignee", 
    "contractor", "created_at", "updated_at", "description", "comment_count"
]

CONTRACTOR_FIELDS = [
    "id", "title", "status", "priority", "site", "created_at"
]

def get_fields_for_role(is_admin: bool) -> List[str]:
    return ADMIN_FIELDS if is_admin else CONTRACTOR_FIELDS

def extract_issue_data(issue: Any, fields: List[str]) -> Dict[str, Any]:
    # Extract site name
    site_name = ""
    if hasattr(issue, 'location') and issue.location and hasattr(issue.location, 'floor_plan') and issue.location.floor_plan:
        site_name = str(issue.location.floor_plan.site_id)  # Best effort if site object not loaded, maybe site_id
        
    # Extract assignee/contractor name
    assignee_name = ""
    if hasattr(issue, 'assignments') and issue.assignments:
        assignee_names = [a.contractor.name for a in issue.assignments if a.contractor]
        assignee_name = ", ".join(assignee_names)

    # Convert enum to string for status/priority
    status = issue.status.value if hasattr(issue.status, 'value') else str(issue.status)
    issue_type = issue.issue_type.value if hasattr(issue.issue_type, 'value') else str(issue.issue_type)
    
    comment_count = len(issue.comments) if hasattr(issue, 'comments') and issue.comments else 0
    
    data = {
        "id": issue.id,
        "title": issue.title,
        "status": status,
        "priority": issue_type,
        "site": site_name,
        "assignee": assignee_name,
        "contractor": assignee_name, # Usually same as assignee in this system
        "created_at": issue.created_at.strftime("%Y-%m-%d %H:%M:%S") if issue.created_at else "",
        "updated_at": issue.updated_at.strftime("%Y-%m-%d %H:%M:%S") if issue.updated_at else "",
        "description": issue.description or "",
        "comment_count": comment_count
    }
    
    return {f: data.get(f, "") for f in fields}

def generate_csv(issues: List[Any], is_admin: bool) -> bytes:
    fields = get_fields_for_role(is_admin)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fields)
    
    writer.writeheader()
    for issue in issues:
        writer.writerow(extract_issue_data(issue, fields))
        
    return output.getvalue().encode('utf-8')

def generate_excel(issues: List[Any], is_admin: bool) -> bytes:
    fields = get_fields_for_role(is_admin)
    wb = Workbook()
    ws = wb.active
    ws.title = "Issues Export"
    
    # Write header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    
    for col_num, field_name in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_num, value=field_name.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        
    # Write data
    for row_num, issue in enumerate(issues, 2):
        data = extract_issue_data(issue, fields)
        for col_num, field_name in enumerate(fields, 1):
            ws.cell(row=row_num, column=col_num, value=data[field_name])
            
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50)
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def generate_pdf(issues: List[Any], is_admin: bool) -> bytes:
    fields = get_fields_for_role(is_admin)
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    elements.append(Paragraph("Space360 Issues Export", styles['Title']))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Generated at: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    elements.append(Spacer(1, 12))
    
    # Table Data
    headers = [f.replace("_", " ").title() for f in fields]
    data = [headers]
    
    for issue in issues:
        row_data = extract_issue_data(issue, fields)
        # Limit text length for PDF table
        row = []
        for f in fields:
            val = str(row_data[f])
            if len(val) > 40:
                val = val[:37] + "..."
            row.append(val)
        data.append(row)
        
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#F2F2F2")),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    return output.getvalue()
